# -*- coding: utf-8 -*-
"""
Created on Wed Aug 29 13:04:16 2018

@author: zmohamadazri
"""
import os
from os.path import join, dirname, abspath, exists
import pandas as pd
import openpyxl
from openpyxl.styles import Font, Border, Side, PatternFill
import smtplib
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
import email.mime.application
import datetime

#parentfolder = abspath(join(masterfolder, os.pardir))
masterfolder = dirname(abspath(__file__))

def mainfile(countrycode, no):
    df = []
    for i in no:
        excel1 = join(join(masterfolder,countrycode+i), 'URL Checking.xlsx')
        
        if exists(excel1):
            df.append(pd.read_excel(excel1))
        else:
            pass
        
    dfmasterfile = pd.concat(df)#,sort=False)
    dfmasterfile = dfmasterfile[['Source','Real URL','URL','STP Name','Key Series','Frequency',
                                 'Level','System ID','Update Method','Remark']]
    
    dfmasterfile = dfmasterfile.rename(columns={'Real URL':'URL1','URL':'URL2_MacroPurpose','STP Name':'STP_Name',
                                                'Key Series':'Key_Status','Level':'Edge_Level','System ID':'System_ID',
                                                'Update Method':'Update_Method'})
    
    allmasterfilefolder = join(masterfolder, 'All Master Files')
    
    if not exists(allmasterfilefolder):
        os.makedirs(allmasterfilefolder)
    else:
        pass
    
    masterfiles = join(allmasterfilefolder, countrycode + ' URL Checking.xlsx')
    
    wb = openpyxl.Workbook(masterfiles)
    wb.save(masterfiles)
    wb1 = openpyxl.load_workbook(masterfiles)
    sheet1 = wb1['Sheet']
    
    for dim1 in [['A',7],['B',20],['C',20],['D',40],['E',11],['F',15],['G',11],['H',10],['I',16],['J',25]]:
        sheet1.column_dimensions[dim1[0]].width = dim1[1]
        
    writer = pd.ExcelWriter(masterfiles,engine='openpyxl')
    writer.book = wb1
    writer.sheets = dict((ws.title,ws) for ws in wb1.worksheets)
    dfmasterfile.to_excel(writer,sheet_name='Sheet',index=False)
    writer.save()
    
    return masterfiles

def consolidatereport(countrycode, no, reportdate):
    df = []
    for i in no:
        excel1 = join(join(masterfolder,countrycode +i+'\\Consolidated Report'),countrycode + ' Consolidated Report.xlsx')
        
        if exists(excel1):
            df.append(pd.read_excel(excel1))
            os.remove(excel1)
        else:
            pass
    
    if not df == []:
        df3 = pd.concat(df)
        df3['Requested Time'] = pd.to_datetime(df3['Requested Time'])
        df3.sort_values(['Requested Time'], ascending=0, inplace=True)  
        df3['Requested Time'] = df3['Requested Time'].dt.strftime('%d-%m-%Y %I:%M:%S %p')
        
        consolidatefolder = join(masterfolder, countrycode + ' Consolidated Report')
        
        if not exists(consolidatefolder):
            os.makedirs(consolidatefolder)
        else:
            pass
        
        consolexcel = join(consolidatefolder , countrycode + ' Consolidated Report ' + reportdate + '.xlsx')
        
        wb = openpyxl.Workbook(consolexcel)
        wb.save(consolexcel)
        wb1 = openpyxl.load_workbook(consolexcel)
        sheet = wb1['Sheet']
        
        writer = pd.ExcelWriter(consolexcel,engine='openpyxl')
        writer.book = wb1
        writer.sheets = dict((ws.title,ws) for ws in wb1.worksheets)
        df3.to_excel(writer,sheet_name='Sheet',index=False)
        writer.save()
        
        for colNum in range(1,13):
            sheet.cell(row=1,column=colNum).fill = PatternFill(fill_type='solid', start_color='B2B2B2', end_color='B2B2B2')
                    
        for dim1 in [['A',8],['B',36],['C',25],['D',25],['E',23],['F',8],['G',11],['H',12],['I',12],['J',8],['K',25],['L',22]]:
            sheet.column_dimensions[dim1[0]].width = dim1[1]
        
        for j in no:
            excel2 = join(join(masterfolder,countrycode +j), 'URL Checking.xlsx')
            df1 = pd.read_excel(excel2)
            
            k = 2
            while sheet.cell(row=k, column=2).value != None:
                for la in df1.index:
                    if sheet.cell(row=k, column=9).value == df1.loc[la, 'System ID']:
                        sheet.cell(row=k, column=2).hyperlink = df1.loc[la, 'Real URL']
                        
                sheet.cell(row=k, column=2).font = Font(color = "0000FF")
                sheet.cell(row=k, column=3).font = Font(color = "FFFF0000")
                    
                for colnum in range(1,13):
                    sheet.cell(row=k,column=colnum).border = Border(top = Side(border_style='thin', color='FF000000'),
                                                                    right = Side(border_style='thin', color='FF000000'),
                                                                    bottom = Side(border_style='thin', color='FF000000'),
                                                                    left = Side(border_style='thin', color='FF000000'))
                
                k += 1
                
            wb1.save(consolexcel)
        
        return consolexcel

def sendconso(fromaddr, toaddr, ccaddr, consolfile, countrycode, reportdate, masterfile):
    serverhost = 'ceicdata-com.mail.protection.outlook.com'
    
    msg = MIMEMultipart('alternative')
    
    msg['From'] = 'RCTCore <RCTCore@isimarkets.com>'
    msg['To'] = toaddr
    msg['CC'] = ccaddr
    toaddrs = [toaddr] + ccaddr.split(',')
    
    if not consolfile == None:
        msg['Subject'] = countrycode+' | Consolidated Report '+ reportdate
        
        text = 'Hi Team,\n\nKindly refer attachment above for '+countrycode+' Release Report on '+reportdate+'.'+'\n\n'+\
        'If you have any enquiry, please do not hesitate to contact us at '+fromaddr+'.\n\nThank you.'
        
        part = MIMEText(text, 'plain')
        msg.attach(part)
        
        fp=open(consolfile,'rb')
        att = email.mime.application.MIMEApplication(fp.read(),_subtype="xlsx")
        fp.close()
        att.add_header('Content-Disposition','attachment',filename=consolfile[1+consolfile.rfind('\\'):])
        msg.attach(att)
        
    else:
        msg['Subject'] = countrycode+' | Consolidated Report '+ reportdate + ' | No Release Detected'
        
        text = 'Hi Team,\n\nFor your information, there are no '+countrycode+' release detected on '+reportdate+'.'+'\n\n'+\
        'If you have any enquiry, please do not hesitate to contact us at RCTCore@isimarkets.com.\n\nThank you.'
        
        part = MIMEText(text, 'plain')
        msg.attach(part)
        
    fp=open(masterfile,'rb')
    attm = email.mime.application.MIMEApplication(fp.read(),_subtype="xlsx")
    fp.close()
    attm.add_header('Content-Disposition','attachment',filename=masterfile[masterfile.rfind('\\'):])
    msg.attach(attm)
    
    server = smtplib.SMTP(serverhost)
    server.sendmail(fromaddr, toaddrs, msg.as_string())
    server.quit()
 
def sendall(*args):
    fromadd = 'youremail@lala.com'
    timenow = (datetime.datetime.now()+datetime.timedelta(hours=8)).strftime('%I %p')
    today = (datetime.datetime.now()+datetime.timedelta(hours=8)).strftime('%A')
    reportsdate = datetime.datetime.now()+datetime.timedelta(hours=8)
    
    for i in args:
        if not today in i['Dont_Send_Days'] and any(timenow == x for x in i['Reports_Time']):
            
            if i['First_Day']+' '+i['Reports_Time'][0] != today+' '+timenow:
                
                for report_time,report_lag in zip(i['Reports_Time'], i['Reports_Lag']):
                    
                    if timenow == report_time:
                        consfile = consolidatereport(i['Country'], i['No'], (reportsdate+datetime.timedelta(days=report_lag)).strftime('%d-%b-%Y'))
                        mtrfile = mainfile(i['Country'], i['No'])
                        sendconso(fromadd,i['To'],i['CC'],consfile,i['Country'], (reportsdate+datetime.timedelta(days=report_lag)).strftime('%d-%b-%Y'), mtrfile)
            
            else:
                consfile = consolidatereport(i['Country'], i['No'], (reportsdate-datetime.timedelta(days=len(i['Dont_Send_Days'])+1)).strftime('%d-%b-%Y'))
                mtrfile = mainfile(i['Country'], i['No'])
                sendconso(fromadd,i['To'],i['CC'],consfile,i['Country'], (reportsdate-datetime.timedelta(days=len(i['Dont_Send_Days'])+1)).strftime('%d-%b-%Y'), mtrfile)
                
            if not consfile == None:
                os.remove(consfile)

if __name__ == '__main__':
    toadd = 'youremail@lala.com'
    ccadd = 'youremail@lala.com, youremail@lala.com'
    
    sendall({'Country':'HKG', 'No':['1','2'], 'To':toadd, 'CC':ccadd, 'Reports_Time':['09 AM','06 PM'], 'Reports_Lag':[-1,0], 'First_Day':'Tuesday',
             'Dont_Send_Days':['Sunday']},
              {'Country':'MAC', 'No':['1','2'], 'To':toadd, 'CC':ccadd, 'Reports_Time':['09 AM','06 PM'], 'Reports_Lag':[-1,0], 'First_Day':'Monday',
               'Dont_Send_Days':['Sunday']},
              {'Country':'PAK', 'No':[''], 'To':toadd, 'CC':ccadd, 'Reports_Time':['02 PM','11 PM'], 'Reports_Lag':[0,0], 'First_Day':'Monday',
               'Dont_Send_Days':['Sunday']},
              {'Country':'PHI', 'No':[''], 'To':toadd, 'CC':ccadd, 'Reports_Time':['12 PM','08 PM'], 'Reports_Lag':[0,0], 'First_Day':'Monday',
               'Dont_Send_Days':['Saturday','Sunday']})
